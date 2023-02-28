import openpyxl

def import_company_sheet(filename):

    # Load workbook
    workbook = openpyxl.load_workbook(filename, data_only=True)

    # Select sheet
    sheet = workbook['Company']

    # Extract data
    header = [cell.value for cell in sheet[1]]
    data = [cell.value for cell in sheet[2]]

    # Create dictionary
    company_dict = {}
    for i, field in enumerate(header):
        company_dict[field] = data[i]

    return company_dict



def import_responsible_sheet(filename):

    # Load workbook
    workbook = openpyxl.load_workbook(filename, data_only=True)

    # Select sheet
    sheet = workbook['Responsible']

    # Extract data
    header = [cell.value for cell in sheet[1]]
    data = [cell.value for cell in sheet[2]]

    # Create dictionary
    responsible_dict = {}
    for i, field in enumerate(header):
        responsible_dict[field] = data[i]

    return responsible_dict


def import_employee_sheet(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Employee"]
    employee_data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        num_inss, name, base_salary, other_remuneration = row
        if num_inss is not None:
            employee_data[num_inss] = {"Nome_do_Trabalhador": name or "",
                                       "Salario_base": base_salary or 0,
                                       "Outras_remuneracoes": other_remuneration or 0}
    return employee_data


def read_excel_data(filename):
    return import_company_sheet(filename), import_responsible_sheet(filename), import_employee_sheet(filename)